import os
import pdfplumber
from openpyxl import load_workbook
import xlrd
from utils.config import logger

class DataHandler:
    def __init__(self):
        pass

    def extract_text_from_pdf(self, file_path):
        try:
            text = ""
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    text += page.extract_text() + "\n"
            return text
        except Exception as e:
            logger.error(f"Error reading PDF file {file_path}: {e}")
            raise

    def fetch_from_files(self, file_path: str):
        try:
            data = []
            # Normalize path
            file_path = os.path.abspath(file_path)

            if os.path.isfile(file_path):
                # Process single file
                if file_path.lower().endswith('.pdf'):
                    # Read PDF file
                    content = self.extract_text_from_pdf(file_path)
                    data.append({"data": content, "source": os.path.basename(file_path)})

                elif file_path.lower().endswith('.xlsx') or file_path.lower().endswith('.xls'):
                    content = self.extract_data_from_excel(file_path)
                    data.append({"data": content, "source": os.path.basename(file_path)})
                else:
                    with open(file_path, 'r') as file:
                        content = file.read()
                        data.append({"data": content, "source": os.path.basename(file_path)})

            print("Successfully fetched data from file")
            return data
        except Exception as e:
            logger.error(f"Error reading files from {file_path}: {e}")
            raise

    def extract_data_from_excel(self, file_path):
        ext = os.path.splitext(file_path)[1].lower()
        data = []

        if ext == '.xlsx':
            # Read .xlsx file using openpyxl
            workbook = load_workbook(filename=file_path, data_only=True)
            for sheet in workbook.sheetnames:
                worksheet = workbook[sheet]
                for row in worksheet.iter_rows(values_only=True):
                    data.append(row)
        elif ext == '.xls':
            # Read .xls file using xlrd
            workbook = xlrd.open_workbook(file_path)
            sheet = workbook.sheet_by_index(0)
            for row_idx in range(sheet.nrows):
                data.append(sheet.row_values(row_idx))

        return data
